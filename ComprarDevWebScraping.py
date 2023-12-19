import base64
import bs4
import requests
import openpyxl


def crear_sopa():
    global link
    pedido = requests.get(link)

    sopa = bs4.BeautifulSoup(pedido.text, 'lxml')
    return sopa


def encontrar_datos(sopa):
    # Listas Datos Finales
    lista_cadenas_codificadas = []
    lista_nombres_productos = []
    lista_precios_amazon = []
    lista_precios_pale = []

    # Buscamos <span> donde esta el código y nombre
    span_elements = sopa.find_all('span', {"class": "ljoptimizer", "data-loc": True})

    # Quitamos la mitad de elementos, pues aparecen repetidos
    for e in range(0, int(len(span_elements)/2)):
        span_elements.pop()

    # Añadimos códigos y nombres a Listas Datos Finales
    for elemento in span_elements:
        lista_cadenas_codificadas.append(elemento["data-loc"])
        lista_nombres_productos.append(elemento.text.strip())

    # Buscamos <td> donde están los precios
    td_elements = sopa.find_all("td")

    # Con posiciones encontramos todos los precios y añadimos a Listas Datos Finales
    contador = 6
    for e in range(0, (len(span_elements))):
        lista_precios_amazon.append(td_elements[contador].text)
        lista_precios_pale.append(td_elements[contador + 1].text)
        contador += 4

    # Devolvemos Datos
    return lista_cadenas_codificadas, lista_nombres_productos, lista_precios_amazon, lista_precios_pale


def modificar_cadenas(cadena):
    indice = cadena.find("%")
    if indice != -1:
        return cadena[:indice]
    else:
        return cadena


def modificar_links(link_amazon):
    indice = link_amazon.find("a")
    if indice != -1:
        return link_amazon[indice:]
    else:
        return link_amazon

def decodificar(cadena):
    # Añade el relleno adecuado ('=') al final de la cadena
    while len(cadena) % 4 != 0:
        cadena += '='

    cadena_decodificada = base64.b64decode(cadena).decode('utf-8')

    return cadena_decodificada


# Pedimos Link y Nombre Archivo
link = input("Introduce el Link")
nombre_archivo = input("¿Qué nombre le quieres poner a tu archivo? (todo junto)")

# Creamos la sopa
sopa = crear_sopa()

# Sacamos Datos de la Web y los asignamos a variables
lista_cadenas, lista_nombres, lista_precios_amazon, lista_precios_pale = encontrar_datos(sopa)

# Modificamos las cadenas codificadas (para quedarnos con el link raiz de amazon)
lista_cadenas_modificada = [modificar_cadenas(cadena) for cadena in lista_cadenas]

# Pasamos las Cadenas a Links y Los Cortamos
lista_links_amazon = [decodificar(cadena) for cadena in lista_cadenas_modificada]
lista_links_amazon = [modificar_links(link_amazon) for link_amazon in lista_links_amazon]

# Creamos Documento y Hoja
workbook = openpyxl.Workbook()
hoja = workbook.active

# Agregamos Nombres Columnas en 1ª fila
nombres_colum = ["Producto", "Link Amazon", "Precio Amazon", "Precio Palé", "Precio Wallapop", "Link Wallapop", "Estimación Venta"]
for num_colum, nombre_colum in enumerate(nombres_colum, 1):
    cell = hoja.cell(row=1, column=num_colum)
    cell.value = nombre_colum

# Metemos todos los datos en una lista
lista_datos_final = [lista_nombres, lista_links_amazon, lista_precios_amazon, lista_precios_pale]

# Pasamos datos a hoja Excel
for num_colum, lista_datos in enumerate(lista_datos_final, 1):
    for num_fila, dato in enumerate(lista_datos, 2):
        cell = hoja.cell(row=num_fila, column=num_colum)
        cell.value = dato

# Guardamos Documento
workbook.save(f"{nombre_archivo}.xlsx")

# Cerramos Documento
workbook.close()
